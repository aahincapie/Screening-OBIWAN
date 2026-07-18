"""
Export layer — CSV bundles, the VM0047 workbook, and a methods note.

Everything a user downloads carries its provenance tier. A screening number that
leaves the app without the label "IPCC defaults, indicative only" attached will be
quoted back as if it were measured, and that is the single most likely way this tool
causes harm.
"""

from __future__ import annotations

import io
import json
import logging
import zipfile
from datetime import datetime, timezone
from typing import Dict, Optional

import pandas as pd

from src.carbon_curve import TIER_LABELS, ResolvedCurve
from src.vm0047 import EngineResults

logger = logging.getLogger(__name__)


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")


# ---------------------------------------------------------------------------
# CSV bundle
# ---------------------------------------------------------------------------

def build_csv_bundle(tables: Dict[str, pd.DataFrame], methods_note: str = "") -> bytes:
    """Zip a set of named DataFrames into one download."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, frame in tables.items():
            if frame is None or frame.empty:
                continue
            zf.writestr(f"{name}.csv", frame.to_csv(index=False))
        if methods_note:
            zf.writestr("METHODS.md", methods_note)
    buffer.seek(0)
    return buffer.getvalue()


def config_json(config_dict: dict) -> bytes:
    """Serialise the full run configuration so a result can be reproduced exactly."""
    payload = {
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "tool": "Screening-OBIWAN",
        "config": config_dict,
    }
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


# ---------------------------------------------------------------------------
# Methods note
# ---------------------------------------------------------------------------

def methods_note(
    curve: ResolvedCurve,
    results: EngineResults,
    aoi_name: str,
    aoi_area_ha: float,
    hansen_asset: str,
    window: tuple[int, int],
    gedi_available: bool,
    gedi_reason: str = "",
) -> str:
    """A Markdown methods statement that travels with every export."""
    setup = results.setup
    summary = results.summary
    tier_label = TIER_LABELS.get(curve.tier, curve.tier)

    caveat = ""
    if curve.is_indicative_only:
        caveat = (
            "\n> **These figures are indicative only.** The carbon curve rests entirely "
            "on published ecological-zone averages with no measurement from this site. "
            "Treat the credit volume as an order-of-magnitude screen, not a projection. "
            "A field inventory or local allometric data would materially change it.\n"
        )

    provenance = "\n".join(f"- {line}" for line in curve.provenance)

    return f"""# Screening-OBIWAN — Methods Note

Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}

## Scope and status

This is an **ex-ante screening** output. It is not a validated VM0047 quantification,
and it is not a substitute for a PDD, a field inventory, or a validation/verification
body's assessment.
{caveat}
## Area of interest

- Source: `{aoi_name}`
- Total AOI area: {aoi_area_ha:,.1f} ha
- Project area after plantable fraction: {setup.area_ha:,.1f} ha

## Forest change

Forest transitions come exclusively from the **Hansen Global Forest Change** dataset
(`{hansen_asset}`), over the window **{window[0]}-{window[1]}**. No national or
regional land-cover product is used, so the classification is globally consistent and
reproducible anywhere on Earth.

Known limitation: Hansen's `gain` band spans 2000-2012 only and carries no year, so
the forest-gain class is used solely to calibrate the regeneration curve and never as
creditable project area.

## Biomass

{"GEDI L4A footprints were sampled and stratified by transition class. Per-stratum "
 "mean AGBD is a design-based estimate with a hybrid standard error combining the "
 "sampling variance of the mean and GEDI's own per-footprint model variance."
 if gedi_available else
 f"GEDI was not used. {gedi_reason}"}

## Carbon accumulation curve

- **Provenance tier: {tier_label}**
- Intervention: {curve.intervention_type}
- Ecological zone: {curve.ecological_zone}
- Resolved AGB ceiling: {curve.ceiling_agb:,.1f} t d.m./ha
- Root:shoot ratio: {curve.effective_root_shoot:.2f}

Parameter provenance:

{provenance}

## Quantification

VM0047 Area-based approach. Deduction chain: baseline removals, project emissions,
leakage, dynamic performance benchmark, uncertainty, non-permanence buffer.

- Uncertainty applied: {100 * setup.uncertainty_pct:.0f}%
  - {setup.uncertainty_note}
- Performance benchmark: {100 * setup.deductions.performance_benchmark_pct:.0f}%
- Non-permanence buffer: {100 * setup.deductions.non_permanence_risk:.0f}%
- Leakage: {100 * setup.deductions.leakage_pct:.0f}%
- Baseline removals: {setup.baseline_rate_tco2e_ha_yr:.3f} tCO2e/ha/yr

## Headline results

| Metric | Value |
|---|---|
| Crediting period | {summary['crediting_period']} years |
| Total gross removals | {summary['total_gross_removals']:,.0f} tCO2e |
| Total net ERs / VCUs | {summary['total_net_ers']:,.0f} tCO2e |
| Average annual ERs | {summary['avg_annual_ers']:,.0f} tCO2e/yr |
| ER per ha per year | {summary['er_per_ha_per_yr']:.3f} tCO2e/ha/yr |
| Total deduction rate | {summary['deduction_pct']:.1f}% |

## Data sources

- Hansen Global Forest Change v1.x — Hansen et al. (2013), *Science* 342:850-853
- GEDI L4A Footprint Level Aboveground Biomass Density — Dubayah et al., ORNL DAAC
- Design-based estimation follows Patterson et al. (2019) and the GEDI L4B ATBD
- VM0047 Afforestation, Reforestation and Revegetation, Verra

Default growth parameters derive from IPCC 2019 Refinement to the 2006 Guidelines,
Vol. 4, Ch. 4. **Verify every default against the current source tables before use in
a submission** — the values shipped with this tool are screening approximations.
"""


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

def build_workbook(
    results: EngineResults,
    curve: ResolvedCurve,
    scenarios: Optional[Dict[str, EngineResults]] = None,
    transition_summary: Optional[pd.DataFrame] = None,
    trend_table: Optional[pd.DataFrame] = None,
) -> bytes:
    """Build the VM0047 screening workbook.

    Sheets: Summary, Project Parameters, Carbon Curve (with chart), ER Projections
    (with chart), Scenario Comparison, Land Cover, GEDI Trend.
    """
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.chart import LineChart, Reference  # noqa: PLC0415
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    setup = results.setup
    summary = results.summary

    wb = Workbook()

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B5E3F")
    title_font = Font(name="Calibri", bold=True, size=14, color="1B5E3F")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_header(ws, row: int, n_cols: int) -> None:
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

    def write_frame(ws, frame: pd.DataFrame, start_row: int = 1) -> int:
        """Write a DataFrame with a styled header. Returns the last row written."""
        for col, name in enumerate(frame.columns, 1):
            ws.cell(start_row, col, str(name))
        write_header(ws, start_row, len(frame.columns))
        for r, (_, record) in enumerate(frame.iterrows(), start_row + 1):
            for c, value in enumerate(record.values, 1):
                cell = ws.cell(r, c, value.item() if hasattr(value, "item") else value)
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.000" if abs(float(value)) < 10 else "#,##0"
        for col in range(1, len(frame.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        return start_row + len(frame)

    # --- Sheet 1: Summary -------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("VM0047 EX-ANTE SCREENING SUMMARY", ""),
        ("", ""),
        ("PROVENANCE", ""),
        ("Carbon curve tier", TIER_LABELS.get(curve.tier, curve.tier)),
        ("Intervention type", curve.intervention_type),
        ("Ecological zone", curve.ecological_zone),
        ("AGB ceiling (t d.m./ha)", round(curve.ceiling_agb, 1)),
        ("", ""),
        ("PROJECT", ""),
        ("Project name", summary["project_name"]),
        ("Project area (ha)", round(summary["total_area_ha"], 1)),
        ("Crediting period (years)", summary["crediting_period"]),
        ("Baseline removals (tCO2e/ha/yr)", round(setup.baseline_rate_tco2e_ha_yr, 4)),
        ("", ""),
        ("EMISSION REDUCTIONS", ""),
        ("Total gross removals (tCO2e)", round(summary["total_gross_removals"])),
        ("Total project emissions (tCO2e)", round(summary["total_project_emissions"])),
        ("Total leakage (tCO2e)", round(summary["total_leakage"])),
        ("Total uncertainty deduction (tCO2e)", round(summary["total_uncertainty_ded"])),
        ("Total PB deduction (tCO2e)", round(summary["total_pb_ded"])),
        ("Total buffer contribution (tCO2e)", round(summary["total_buffer"])),
        ("TOTAL NET ERs / VCUs (tCO2e)", round(summary["total_net_ers"])),
        ("", ""),
        ("PERFORMANCE", ""),
        ("Average annual ERs (tCO2e/yr)", round(summary["avg_annual_ers"])),
        ("Peak annual ERs (tCO2e)", round(summary["peak_annual_ers"])),
        ("Peak year", summary["peak_year"]),
        ("ER per ha per year", round(summary["er_per_ha_per_yr"], 4)),
        ("Total deduction rate (%)", round(summary["deduction_pct"], 1)),
        ("", ""),
        ("UNCERTAINTY", ""),
        ("Applied uncertainty (%)", round(100 * setup.uncertainty_pct, 1)),
        ("Rationale", setup.uncertainty_note),
    ]
    if curve.is_indicative_only:
        rows += [
            ("", ""),
            ("WARNING", "Curve uses IPCC defaults only — figures are indicative, "
                        "not a projection. Obtain site measurement before relying on them."),
        ]
    for r, (label, value) in enumerate(rows, 1):
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
    ws.cell(1, 1).font = title_font
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60
    ws["B33"].alignment = Alignment(wrap_text=True, vertical="top")

    # --- Sheet 2: Project Parameters -------------------------------------
    ws = wb.create_sheet("Project Parameters")
    param_rows = [
        ("CURVE PARAMETERS", "", ""),
        ("Provenance tier", TIER_LABELS.get(curve.tier, curve.tier), ""),
        ("Root:shoot ratio", round(curve.effective_root_shoot, 3), "Below-ground / above-ground"),
        ("Growth multiplier", round(curve.modifiers.growth_multiplier, 3), "Combined site modifiers"),
        ("Ceiling multiplier", round(curve.modifiers.ceiling_multiplier, 3), ""),
        ("", "", ""),
        ("PROVENANCE TRAIL", "", ""),
    ]
    param_rows += [("", line, "") for line in curve.provenance]
    param_rows += [
        ("", "", ""),
        ("SPECIES MIX", "Area fraction", "Growth model"),
    ]
    param_rows += [
        (s.name, round(s.area_fraction, 3), s.growth_model) for s in curve.species
    ]
    param_rows += [
        ("", "", ""),
        ("DEDUCTIONS (VM0047 Eq. 30)", "Value", "Note"),
        ("Uncertainty", round(setup.uncertainty_pct, 4), setup.uncertainty_note),
        ("Performance benchmark", setup.deductions.performance_benchmark_pct, "Dynamic PB"),
        ("Non-permanence risk", setup.deductions.non_permanence_risk, "Buffer pool"),
        ("Leakage", setup.deductions.leakage_pct, "Activity shifting"),
    ]
    for r, values in enumerate(param_rows, 1):
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
    ws.cell(1, 1).font = title_font
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 64
    ws.column_dimensions["C"].width = 40

    # --- Sheet 3: Carbon Curve (+ chart) ---------------------------------
    ws = wb.create_sheet("Carbon Curve")
    curve_frame = curve.to_frame(setup.project.crediting_period_years)
    last = write_frame(ws, curve_frame.round(4))

    chart = LineChart()
    chart.title = "Carbon accumulation curve (tCO2e/ha)"
    chart.y_axis.title = "tCO2e / ha"
    chart.x_axis.title = "Stand age (years)"
    chart.width, chart.height = 24, 13
    chart.add_data(Reference(ws, min_col=5, min_row=1, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
    ws.add_chart(chart, "H2")

    # --- Sheet 4: ER Projections (+ chart) -------------------------------
    ws = wb.create_sheet("ER Projections")
    er_frame = results.annual.round(3)
    last = write_frame(ws, er_frame)

    net_col = list(er_frame.columns).index("net_ers_tco2e") + 1
    year_col = list(er_frame.columns).index("calendar_year") + 1
    chart = LineChart()
    chart.title = "Annual net ERs (tCO2e)"
    chart.y_axis.title = "tCO2e"
    chart.x_axis.title = "Calendar year"
    chart.width, chart.height = 24, 13
    chart.add_data(Reference(ws, min_col=net_col, min_row=1, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=year_col, min_row=2, max_row=last))
    ws.add_chart(chart, f"{get_column_letter(len(er_frame.columns) + 2)}2")

    # --- Sheet 5: Scenario Comparison ------------------------------------
    if scenarios:
        from src.vm0047 import scenario_comparison  # noqa: PLC0415

        ws = wb.create_sheet("Scenario Comparison")
        write_frame(ws, scenario_comparison(scenarios))
        ws.column_dimensions["A"].width = 34

    # --- Sheet 6: Land Cover ---------------------------------------------
    if transition_summary is not None and not transition_summary.empty:
        ws = wb.create_sheet("Land Cover")
        write_frame(ws, transition_summary)

    # --- Sheet 7: GEDI Trend ---------------------------------------------
    if trend_table is not None and not trend_table.empty:
        ws = wb.create_sheet("GEDI Trend")
        write_frame(ws, trend_table.round(4))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def workbook_filename(project_name: str) -> str:
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in project_name)[:40]
    return f"VM0047_Screening_{safe}_{_timestamp()}.xlsx"


def bundle_filename(project_name: str) -> str:
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in project_name)[:40]
    return f"ScreeningOBIWAN_{safe}_{_timestamp()}.zip"
